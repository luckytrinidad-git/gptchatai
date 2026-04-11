import fitz  # PyMuPDF
import json
import io
from PIL import Image
import easyocr
import openpyxl
import numpy as np

# initialize once (IMPORTANT for performance)
reader = easyocr.Reader(['en'], gpu=False)

def extract_text(file_name, file_bytes):

    filename = file_name.lower()

    # =========================
    # PDF
    # =========================
    if filename.endswith(".pdf"):

        text = ""
        pdf = fitz.open(stream=file_bytes, filetype="pdf")

        for page in pdf:

            # =========================
            # 1. FAST TEXT EXTRACTION
            # =========================
            page_text = page.get_text("text")

            if page_text and page_text.strip():
                text += page_text + "\n"
                continue

            # =========================
            # 2. OCR FALLBACK (NO POPPLER, NO TESSERACT DEPENDENCY)
            # =========================
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # upscale for better OCR

            img = Image.frombytes(
                "RGB",
                [pix.width, pix.height],
                pix.samples
            )
            
            img_np = np.array(img)

            result = reader.readtext(img_np, detail=0)

            text += " ".join(result) + "\n"

        return text.strip()
    # =========================
    # TXT / CSV
    # =========================
    elif filename.endswith((".txt", ".csv")):
        return file_bytes.decode("utf-8", errors="ignore")

    # =========================
    # JSON
    # =========================
    elif filename.endswith(".json"):
        try:
            data = json.loads(file_bytes.decode("utf-8", errors="ignore"))
            return json.dumps(data, indent=2)
        except:
            return file_bytes.decode("utf-8", errors="ignore")

    # =========================
    # DOCX
    # =========================
    elif filename.endswith(".docx"):
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join([p.text for p in doc.paragraphs])

    # =========================
    # XLSX
    # =========================
    elif filename.endswith(".xlsx"):

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        all_text = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            all_text.append(f"\n===== SHEET: {sheet_name} =====\n")

            for row in sheet.iter_rows(values_only=True):

                cleaned_row = []

                for cell in row:
                    if cell is None:
                        continue

                    # avoid Excel weird objects
                    if isinstance(cell, (int, float, str)):
                        cleaned_row.append(str(cell))
                    else:
                        cleaned_row.append(str(cell))

                if cleaned_row:
                    all_text.append(" | ".join(cleaned_row))

        return "\n".join(all_text).strip()

    # =========================
    # FALLBACK
    # =========================
    return file_bytes.decode("utf-8", errors="ignore")